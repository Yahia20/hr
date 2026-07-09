"""End-to-end tests for the attendance feature (clock in/out, geofence,
office CRUD, WebAuthn ceremony error paths, listing scope, pagination, export).

The real WebAuthn success path needs a physical authenticator, so biometric
verification is exercised up to the ceremony/registration boundary only."""
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from conftest import csrf

RIYADH = (24.7136, 46.6753)
CAIRO = (30.0444, 31.2357)


def _admin_row(admin, email):
    body = admin.get(f"/api/attendance?employee={email}").json()
    return body["rows"][0] if body["rows"] else None


def test_me_initial(new_employee):
    emp, _ = new_employee()
    me = emp.get("/api/attendance/me").json()
    assert me["today"] is None
    assert me["has_credential"] is False
    assert me["require_biometric"] is False
    assert me["require_geofence"] is False  # no offices configured


def test_clock_in_out_flow(new_employee):
    emp, email = new_employee()
    # No offices -> clocking allowed from anywhere.
    r = emp.post("/api/attendance/clock-in", json={"lat": 30.0, "lng": 31.0, "accuracy": 10}, headers=csrf(emp))
    assert r.status_code == 201, r.text
    assert r.json()["clock_in_at"] and r.json()["clock_out_at"] is None

    # A second clock-in the same day is rejected.
    r = emp.post("/api/attendance/clock-in", json={"lat": 30.0, "lng": 31.0}, headers=csrf(emp))
    assert r.status_code == 409 and r.json()["detail"] == "already_clocked_in"

    r = emp.post("/api/attendance/clock-out", json={"lat": 30.0, "lng": 31.0, "accuracy": 5}, headers=csrf(emp))
    assert r.status_code == 200 and r.json()["clock_out_at"]

    body = emp.get("/api/attendance").json()
    assert body["total"] == 1
    row = body["rows"][0]
    assert row["email"] == email and row["worked_hours"] is not None


def test_clock_out_without_clock_in(new_employee):
    emp, _ = new_employee()
    r = emp.post("/api/attendance/clock-out", json={"lat": 1, "lng": 1}, headers=csrf(emp))
    assert r.status_code == 409 and r.json()["detail"] == "not_clocked_in"


def test_geofence_enforced(admin, new_employee):
    r = admin.post(
        "/api/attendance/offices",
        json={"name": "HQ", "lat": RIYADH[0], "lng": RIYADH[1], "radius_m": 150},
        headers=csrf(admin),
    )
    assert r.status_code == 201, r.text
    emp, _ = new_employee()

    # Outside the fence -> 403 with the distance to the nearest office.
    r = emp.post("/api/attendance/clock-in", json={"lat": 25.0, "lng": 47.0, "accuracy": 10}, headers=csrf(emp))
    assert r.status_code == 403 and r.json()["detail"].startswith("outside_geofence:")

    # Missing location while geofencing is enforced -> 400.
    r = emp.post("/api/attendance/clock-in", json={}, headers=csrf(emp))
    assert r.status_code == 400 and r.json()["detail"] == "location_required"

    # Inside the fence -> ok, office name recorded.
    r = emp.post("/api/attendance/clock-in", json={"lat": RIYADH[0] + 0.0005, "lng": RIYADH[1], "accuracy": 20}, headers=csrf(emp))
    assert r.status_code == 201, r.text
    assert r.json()["clock_in_office"] == "HQ"

    me = emp.get("/api/attendance/me").json()
    assert me["require_geofence"] is True


def test_biometric_required_blocks_unregistered(monkeypatch, new_employee):
    monkeypatch.setattr("app.routers.attendance.REQUIRE_BIOMETRIC", True)
    emp, _ = new_employee()
    r = emp.post("/api/attendance/clock-in", json={"lat": 1, "lng": 1}, headers=csrf(emp))
    assert r.status_code == 403 and r.json()["detail"] == "fingerprint_not_registered"


def test_list_scope_and_pagination(admin):
    marker = "PGMARK"
    emps = []
    for i in range(3):
        _seq_email = f"pg{i}@test.com"
        admin.post(
            "/api/auth/users",
            json={"email": _seq_email, "name": f"{marker} {i}", "role": "employee", "department": "Pag", "password": "password123"},
            headers=csrf(admin),
        )
        from conftest import login
        c = login(_seq_email, "password123")
        c.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10}, headers=csrf(c))
        emps.append(c)

    # Filter by the unique marker so the total is deterministic despite other tests.
    body = admin.get(f"/api/attendance?employee={marker}&limit=2&offset=0").json()
    assert body["total"] == 3
    assert len(body["rows"]) == 2
    body2 = admin.get(f"/api/attendance?employee={marker}&limit=2&offset=2").json()
    assert len(body2["rows"]) == 1

    # An employee only sees their own record.
    self_rows = emps[0].get("/api/attendance").json()
    assert self_rows["total"] == 1

    # limit is capped, and a bad date is rejected.
    assert admin.get("/api/attendance?limit=99999").status_code == 422
    assert admin.get("/api/attendance?date_from=nonsense").status_code == 400


def test_office_endpoints_role_guarded(new_employee):
    emp, _ = new_employee()
    assert emp.get("/api/attendance/offices").status_code == 403
    assert emp.post("/api/attendance/offices", json={"name": "x", "lat": 0, "lng": 0}, headers=csrf(emp)).status_code == 403
    assert emp.get("/api/attendance/export").status_code == 403


def test_webauthn_error_paths(new_employee):
    emp, _ = new_employee()
    opts = emp.post("/api/attendance/webauthn/register/begin", headers=csrf(emp)).json()
    assert opts["challenge"] and opts["rp"]["id"]
    assert opts["authenticatorSelection"]["userVerification"] == "required"

    r = emp.post("/api/attendance/webauthn/register/complete", json={"credential": {"id": "x"}}, headers=csrf(emp))
    assert r.status_code == 400 and r.json()["detail"] == "registration_failed"

    r = emp.post("/api/attendance/webauthn/clock/begin", headers=csrf(emp))
    assert r.status_code == 403 and r.json()["detail"] == "fingerprint_not_registered"

    assert emp.get("/api/attendance/webauthn/credentials").json() == []


def test_export_content_type(admin, new_employee):
    emp, _ = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 1, "lng": 1}, headers=csrf(emp))
    r = admin.get("/api/attendance/export")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]


def test_auth_required():
    anon = TestClient(app)
    assert anon.get("/api/attendance").status_code == 401
    assert anon.post("/api/attendance/clock-in", json={}).status_code == 401


def test_risk_zero_accuracy_flagged(admin, new_employee):
    emp, email = new_employee()
    r = emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 0}, headers=csrf(emp))
    assert r.status_code == 201
    row = _admin_row(admin, email)
    assert row["risk"]["level"] == "high"
    assert "zero_accuracy" in row["risk"]["reasons"]


def test_risk_impossible_travel_flagged(admin, new_employee):
    emp, email = new_employee()
    # Clock in in Riyadh and out in Cairo seconds later — physically impossible.
    emp.post("/api/attendance/clock-in", json={"lat": RIYADH[0], "lng": RIYADH[1], "accuracy": 20}, headers=csrf(emp))
    emp.post("/api/attendance/clock-out", json={"lat": CAIRO[0], "lng": CAIRO[1], "accuracy": 20}, headers=csrf(emp))
    row = _admin_row(admin, email)
    assert row["risk"]["level"] == "high"
    assert "impossible_travel" in row["risk"]["reasons"]


def test_risk_clean_when_plausible(admin, new_employee):
    emp, email = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 22}, headers=csrf(emp))
    row = _admin_row(admin, email)
    assert row["risk"]["level"] == "none" and row["risk"]["reasons"] == []


def _hq(admin, radius=150):
    admin.post(
        "/api/attendance/offices",
        json={"name": "HQ", "lat": RIYADH[0], "lng": RIYADH[1], "radius_m": radius},
        headers=csrf(admin),
    )


def test_slack_cannot_widen_fence_far(admin, new_employee):
    # ~240m from a 150m office claiming accuracy=100: slack is capped at 50, so
    # the effective fence is 200m and the punch must be rejected, not accepted.
    _hq(admin)
    emp, _ = new_employee()
    r = emp.post(
        "/api/attendance/clock-in",
        json={"lat": RIYADH[0] + 0.00216, "lng": RIYADH[1], "accuracy": 100},
        headers=csrf(emp),
    )
    assert r.status_code == 403 and r.json()["detail"].startswith("outside_geofence:")


def test_edge_of_fence_is_flagged(admin, new_employee):
    # Just outside the 150m radius but inside radius+slack -> accepted BUT flagged.
    _hq(admin)
    emp, email = new_employee()
    r = emp.post(
        "/api/attendance/clock-in",
        json={"lat": RIYADH[0] + 0.00162, "lng": RIYADH[1], "accuracy": 40},  # ~180m out
        headers=csrf(emp),
    )
    assert r.status_code == 201, r.text
    row = _admin_row(admin, email)
    assert "edge_of_fence" in row["risk"]["reasons"]
    assert row["risk"]["level"] in ("low", "high")


def test_solidly_inside_is_not_flagged(admin, new_employee):
    _hq(admin)
    emp, email = new_employee()
    r = emp.post(
        "/api/attendance/clock-in",
        json={"lat": RIYADH[0] + 0.0003, "lng": RIYADH[1], "accuracy": 20},  # ~33m, well inside
        headers=csrf(emp),
    )
    assert r.status_code == 201
    row = _admin_row(admin, email)
    assert "edge_of_fence" not in row["risk"]["reasons"]


def test_low_precision_flagged(admin, new_employee):
    emp, email = new_employee()  # no offices -> geofence not enforced
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 80}, headers=csrf(emp))
    row = _admin_row(admin, email)
    assert "low_precision" in row["risk"]["reasons"]


def _insert_attendance(user_id, work_date, coords=True):
    from app.db import db
    with db() as conn:
        conn.execute(
            """INSERT INTO attendance
               (user_id, work_date, clock_in_at, clock_in_lat, clock_in_lng, clock_in_accuracy,
                clock_in_office, clock_in_distance_m, clock_in_verified, clock_in_ip)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (user_id, work_date, f"{work_date} 08:00:00",
             24.7 if coords else None, 46.6 if coords else None, 15 if coords else None,
             "HQ", 40.0, 1, "203.0.113.9"),
        )


def test_retention_purges_old_coords_keeps_facts(admin, new_employee):
    from app.db import db
    from app.routers.attendance import purge_location_data
    emp, email = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 24.7, "lng": 46.6, "accuracy": 15}, headers=csrf(emp))
    user_id = _admin_row(admin, email)["user_id"]

    # A punch from ~200 days ago, with coordinates + IP.
    _insert_attendance(user_id, "2020-01-01")
    scrubbed = purge_location_data(90)
    assert scrubbed >= 1

    with db() as conn:
        old = dict(conn.execute(
            "SELECT * FROM attendance WHERE user_id = ? AND work_date = '2020-01-01'", (user_id,)
        ).fetchone())
    # Raw location + IP gone…
    assert old["clock_in_lat"] is None and old["clock_in_lng"] is None
    assert old["clock_in_accuracy"] is None and old["clock_in_ip"] == ""
    # …but the derived audit facts remain.
    assert old["clock_in_office"] == "HQ" and old["clock_in_distance_m"] == 40.0
    assert old["clock_in_verified"] == 1 and old["clock_in_at"] == "2020-01-01 08:00:00"


def test_retention_keeps_recent_rows(admin, new_employee):
    from app.db import db
    from app.routers.attendance import purge_location_data
    emp, email = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 24.7, "lng": 46.6, "accuracy": 15}, headers=csrf(emp))
    row = _admin_row(admin, email)
    purge_location_data(90)
    fresh = _admin_row(admin, email)
    assert fresh["clock_in_lat"] is not None  # today's punch is untouched


def test_retention_disabled_with_zero():
    from app.routers.attendance import purge_location_data
    assert purge_location_data(0) == 0


def test_risk_hidden_from_employee_and_ip_not_leaked(new_employee):
    emp, _ = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 0}, headers=csrf(emp))
    row = emp.get("/api/attendance").json()["rows"][0]
    assert "risk" not in row  # spoof flags not revealed to the employee
    assert "clock_in_ip" not in row  # IP is audit-only, never in the list UI


def test_export_has_risk_and_ip_columns(admin, new_employee):
    emp, _ = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 5}, headers=csrf(emp))
    r = admin.get("/api/attendance/export")
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    headers = [c.value for c in ws[1]]
    for col in ("Risk", "Risk Reasons", "IP (In)", "IP (Out)"):
        assert col in headers


# --- Office networks (Wi-Fi egress IP allow-list) — advisory only -----------

def test_office_networks_crud_and_guards(admin, new_employee):
    # Manager can add; a bare IP is normalised to a /32.
    r = admin.post("/api/attendance/networks", json={"label": "HQ Wi-Fi", "cidr": "203.0.113.7"}, headers=csrf(admin))
    assert r.status_code == 201, r.text
    assert r.json()["cidr"] == "203.0.113.7/32"
    nid = r.json()["id"]

    # A CIDR range is accepted; garbage is rejected by the schema.
    assert admin.post("/api/attendance/networks", json={"cidr": "10.0.0.0/8"}, headers=csrf(admin)).status_code == 201
    assert admin.post("/api/attendance/networks", json={"cidr": "not-an-ip"}, headers=csrf(admin)).status_code == 422

    # Officer reads but can't write or delete; employee can't even read.
    officer, _ = new_employee(role="hr_officer")
    assert officer.get("/api/attendance/networks").status_code == 200
    assert officer.post("/api/attendance/networks", json={"cidr": "192.0.2.0/24"}, headers=csrf(officer)).status_code == 403
    assert officer.delete(f"/api/attendance/networks/{nid}", headers=csrf(officer)).status_code == 403
    emp, _ = new_employee(role="employee")
    assert emp.get("/api/attendance/networks").status_code == 403

    assert admin.delete(f"/api/attendance/networks/{nid}", headers=csrf(admin)).status_code == 204


def test_networks_require_auth():
    anon = TestClient(app)
    assert anon.get("/api/attendance/networks").status_code == 401


def test_ip_on_network_helper(admin):
    from app.routers.attendance import _ip_on_network
    # No networks configured -> feature off (None), nothing to flag.
    assert _ip_on_network("203.0.113.5") is None
    admin.post("/api/attendance/networks", json={"label": "HQ", "cidr": "203.0.113.0/24"}, headers=csrf(admin))
    assert _ip_on_network("203.0.113.5") == 1    # inside the range
    assert _ip_on_network("198.51.100.9") == 0   # outside
    assert _ip_on_network("not-an-ip") == 0      # unparseable can't be on-network


def test_my_ip_endpoint(admin, new_employee):
    assert "ip" in admin.get("/api/attendance/my-ip").json()
    emp, _ = new_employee(role="employee")
    assert emp.get("/api/attendance/my-ip").status_code == 403  # HR-staff only


def test_punch_flagged_when_off_office_network(admin, new_employee):
    # With a network configured, a punch from an IP outside it is flagged
    # off_network (advisory) but still succeeds — the check never blocks.
    admin.post("/api/attendance/networks", json={"label": "HQ", "cidr": "203.0.113.0/24"}, headers=csrf(admin))
    emp, email = new_employee()
    r = emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 5}, headers=csrf(emp))
    assert r.status_code == 201, r.text  # accepted, not blocked

    row = _admin_row(admin, email)
    assert row["clock_in_on_network"] == 0
    assert "off_network" in row["risk"]["reasons"]


def test_no_networks_means_no_flag(admin, new_employee):
    emp, email = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 5}, headers=csrf(emp))
    row = _admin_row(admin, email)
    assert row["clock_in_on_network"] is None  # feature off
    assert "off_network" not in row["risk"]["reasons"]


def test_on_network_hidden_from_employee(admin, new_employee):
    admin.post("/api/attendance/networks", json={"cidr": "203.0.113.0/24"}, headers=csrf(admin))
    emp, _ = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10}, headers=csrf(emp))
    row = emp.get("/api/attendance").json()["rows"][0]
    assert "clock_in_on_network" not in row  # advisory flag is HR-only in the list


def test_export_has_network_columns(admin, new_employee):
    emp, _ = new_employee()
    emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 5}, headers=csrf(emp))
    ws = load_workbook(io.BytesIO(admin.get("/api/attendance/export").content)).active
    headers = [c.value for c in ws[1]]
    for col in ("On Network (In)", "On Network (Out)"):
        assert col in headers


def test_punch_on_network_via_http(admin, new_employee, monkeypatch):
    # The full HTTP clock path with a request IP inside the office network:
    # both punches record on_network=1 and nothing is flagged. (TestClient's own
    # IP can't match a real CIDR, so the on-network path is driven by patching
    # the resolved client IP — the same value uvicorn derives via --proxy-headers.)
    admin.post("/api/attendance/networks", json={"label": "HQ", "cidr": "203.0.113.0/24"}, headers=csrf(admin))
    monkeypatch.setattr("app.routers.attendance.client_ip", lambda request: "203.0.113.50")
    emp, email = new_employee()

    assert emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10, "accuracy": 5}, headers=csrf(emp)).status_code == 201
    assert emp.post("/api/attendance/clock-out", json={"lat": 10, "lng": 10, "accuracy": 5}, headers=csrf(emp)).status_code == 200

    row = _admin_row(admin, email)
    assert row["clock_in_on_network"] == 1 and row["clock_out_on_network"] == 1
    assert "off_network" not in row["risk"]["reasons"]


def test_punch_off_network_via_http_still_succeeds(admin, new_employee, monkeypatch):
    # A controlled outside IP: flagged off_network but the punch is accepted
    # (advisory, never blocks) — the full HTTP path, not just the helper.
    admin.post("/api/attendance/networks", json={"cidr": "203.0.113.0/24"}, headers=csrf(admin))
    monkeypatch.setattr("app.routers.attendance.client_ip", lambda request: "8.8.8.8")
    emp, email = new_employee()
    assert emp.post("/api/attendance/clock-in", json={"lat": 10, "lng": 10}, headers=csrf(emp)).status_code == 201

    row = _admin_row(admin, email)
    assert row["clock_in_on_network"] == 0
    assert "off_network" in row["risk"]["reasons"]


def test_ip_on_network_ipv6_and_multiple_ranges(admin):
    from app.routers.attendance import _ip_on_network
    admin.post("/api/attendance/networks", json={"cidr": "203.0.113.7"}, headers=csrf(admin))      # bare IP -> /32
    admin.post("/api/attendance/networks", json={"cidr": "198.51.100.0/24"}, headers=csrf(admin))  # a v4 range
    admin.post("/api/attendance/networks", json={"cidr": "2001:db8::/32"}, headers=csrf(admin))     # a v6 range
    assert _ip_on_network("203.0.113.7") == 1      # exact /32 match
    assert _ip_on_network("203.0.113.8") == 0      # just outside the /32
    assert _ip_on_network("198.51.100.42") == 1    # inside the v4 range
    assert _ip_on_network("2001:db8::1") == 1      # inside the v6 range
    assert _ip_on_network("2001:dead::1") == 0     # outside the v6 range
