from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ...db.session import get_db
from ...models.violation import Violation
from ...api.deps import get_current_user

router = APIRouter()

_PENALTY_COLORS = {
    "Yellow": "FFFDE7",
    "Orange": "FFF3E0",
    "Red": "FFEBEE",
    "Black": "F3F4F6",
    "Investigation": "F5F3FF",
}


@router.get("/export")
def export_violations(
    employee_name: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    penalty_level: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    query = db.query(Violation)
    if employee_name:
        query = query.filter(Violation.employee_name == employee_name)
    if penalty_level:
        query = query.filter(Violation.penalty_color == penalty_level)
    if date_from:
        try:
            query = query.filter(Violation.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Violation.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            pass

    violations = query.order_by(Violation.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations Report"

    headers = ["ID", "Employee", "Category", "Incident", "Penalty", "Deduction Days",
               "Submitted By", "Comment", "Date"]
    header_fill = PatternFill(start_color="2FB89E", end_color="2FB89E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, v in enumerate(violations, 2):
        ws.append([
            v.id,
            v.employee_name,
            v.category,
            v.incident,
            v.penalty_color,
            v.deduction_days,
            v.submitted_by,
            v.comment,
            v.created_at.strftime("%Y-%m-%d %H:%M") if v.created_at else "",
        ])
        color = _PENALTY_COLORS.get(v.penalty_color, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row, column=5).fill = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or "")) for cell in col
        ) + 4

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=violations_report.xlsx"},
    )
